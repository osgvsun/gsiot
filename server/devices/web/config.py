# -*- coding:utf-8 -*-
import sys,os,platform,threading,json,hashlib
from datetime import datetime
if __name__=='__main__':
    path=sys.path[0]
    sys.path.append(path+"/../..")
from lib import *
from lib.net import *
from app.devices import App
from lib.file.jsonfile import gsJsonFile
app=App(sys.path[0]+"/../..")
route=app.webroute
@route("/agent/<ip>/hi")
def setConfig(route,ip,request):
    return {"result":True}
@route("/api/config/<filename>")
def setConfig(route,filename,request):
    if request.request_data!={}:
        file=gsJsonFile(app.rootpath+"/etc/conf/"+filename)
        file.data=edict(request.request_data)
        file.Savefile()
        app.logservice("info",request.request_data)
        return {"result":True}
    ret={"filename":app.rootpath+"/etc/conf/"+filename,"isexists":os.path.isfile(app.rootpath+"/etc/conf/"+filename)}
    if ret["isexists"]:ret=json.load(open(app.rootpath+"/etc/conf/"+filename))
    return ret
@route.get("/api/log/<logtype>")
def getloglist(route,logtype,request):
    path="{}/etc/log/{}".format(app.rootpath,logtype)
    if logtype=="attlog.txt":
        if not os.path.exists(path):
            os.system("touch {}".format(path))
        return {"result":json.loads("["+open(path,"r").read().replace("\x00","")[1:].encode("utf-8")+"]")}
    elif os.path.exists(path):
        filelist=os.popen("ls {}/*.log".format(path)).read().replace("{}/".format(path),"").split("\n")
        filelist.remove("")
        return {"length":len(filelist),"files":filelist,"folder":path}
    else:return {"result":False}
@route.get("/api/log/md/<logtype>/<filename>")
def getlog(route,logtype,filename,request):
    f=open(app.rootpath+"/etc/log/{}/{}".format(logtype,filename),"r")
    data=msg=f.read().replace("\x00","")[:-1].replace("\n",'\n- ')
    f.close()
    request.response_head['Content-Type'] = CONTENTTYPE[".html"]
    mdfile=open(app.rootpath+"/app/devices/web/markdown.txt","r")
    data=mdfile.read().replace("<%=filedata%>","- "+msg)
    return data
@route.get("/api/log/<logtype>/<filename>")
def getlog(route,logtype,filename,request):
    f=open(app.rootpath+"/etc/log/{}/{}".format(logtype,filename),"r")
    data=f.read().replace("\x00","")
    f.close()
    request.response_head['Content-Type'] = CONTENTTYPE[".html"]
    return data
@route.get("/api/runtime/processlist")
def getProc(route,request):return {"result":System.ProcList()}
@route.get("/api/runtime/process/<find>")
def getProc(route,find,request):return {"result":System.ProcList(find)}
@route.get("/api/runtime/network")
def getNetwork(route,request):return app.net.toDict()
@route.get("/api/runtime/network/<adapter>")
def getNetwork(route,adapter,request):return app.net(adapter)
@route.get("/api/runtime/network/<adapter>/scanap")
def getNetwork(route,adapter,request):return app.net.nic[adapter].ScanAP()
@route.get("/api/runtime/network/<adapter>/<ip>/<mask>/<gw>")
def getNetwork(route,adapter,ip,mask,gw,request):
    lan=app.net.nic[adapter]
    flag=False or lan.data.ip!=ip or lan.data.netmask!=mask or lan.data.gateway!=gw
    if flag and app.net.checkip(ip):
        lan.ChangeIP(ip=ip,netmask=mask,gw=gw)
        lan.data.dhcp=False
        return {"result":True,"cfg":lan.data}
    else:return {"result":False,"msg":"当前信息不全或地址重复"}
@route.get("/api/runtime/network/save")
def setNetwork(route,request):
    data=app.net.setConfigFile()
    app.logservice("info","make /etc/network/interfaces",data)
    f=open("/etc/network/interfaces","w")
    f.write(data)
    f.close()
    return {"result":True}
@route.get("/api/runtime/network/<adapter>/<ssid>/<psk>")
def getNetwork(route,adapter,ssid,psk,request):
    lan=app.net.nic[adapter]
    lan.changeWPConnect(ssid=ssid,psk=psk)
    time.sleep(3)
    lan.iwConfig()
    return {"result":True,"cfg":lan.data}
@route.get("/api/runtime/network/<adapter>/dhcp")
def setNetwork(route,adapter,request):
    lan=app.net.nic[adapter]
    lan.data.dhcp=True
    lan.reDhcp()
    time.sleep(1)
    lan()
    return {"result":True,"cfg":lan.data}
@route.get("/api/runtime/bindnetwork/<adapter>")
def setNetwork(route,adapter,request):
    if adapter in app.net.nic:
        # app.net=Net()
        app.cfg.device.network=adapter
        app.cfg.device.ip=app.net(adapter).ip
        app.cfg.Savefile()
    return {"result":True,"cfg":app.cfg.data}
@route.get("/api/setAttendance")
def setAttendance(route,request):
    try:
        try:from openpyxl import  Workbook 
        except:
            os.system("pip install openpyxl")
            from openpyxl import  Workbook 

        os.system("rm -rf {}/etc/log/attlog.xlsx".format(app.rootpath))
        f=open(app.rootpath+"/etc/log/attlog.txt","r")
        wb = Workbook()
        ws=wb["Sheet"]
        data=json.loads("["+f.read()[1:]+"]")
        f.close()
        if len(data)!=0:
            cell=1
            if len(data)!=0:
                ws["A"+str(cell)]="工号"
                ws["B"+str(cell)]="姓名"
                ws["C"+str(cell)]="卡号"
                ws["D"+str(cell)]="时间"
            for item in data:
                record=edict(item)
                cell+=1
                ws["A"+str(cell)]=record.username
                ws["B"+str(cell)]=record.cname
                ws["C"+str(cell)]=record.cardnumber if "cardnumber" in record else ""
                ws["D"+str(cell)]=record.datetime
            wb.save(app.rootpath+"/etc/log/attlog.xlsx")
            return {"result":True,"code":len(data)}
        return {"result":True,"code":0}
    except:return {"result":False}